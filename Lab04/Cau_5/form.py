from tkinter import *
from openpyxl import *

# open the existing excel file
wb = load_workbook('data.xlsx')

# create the sheet object, which is the current excel sheet
sheet = wb.active

#  region Format the Excel file


def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = 'Name'
    sheet.cell(row=1, column=2).value = 'Course'
    sheet.cell(row=1, column=3).value = 'Semester'
    sheet.cell(row=1, column=4).value = 'Form Number'
    sheet.cell(row=1, column=5).value = 'Contact Number'
    sheet.cell(row=1, column=6).value = 'Email ID'
    sheet.cell(row=1, column=7).value = 'Address'
# endregion

# region set Focuses


def focus1(event):
    course_field.focus_set()


def focus2(event):
    sem_field.focus_set()


def focus3(event):
    form_no_field.focus_set()


def focus4(event):
    contact_no_field.focus_set()


def focus5(event):
    email_field.focus_set()


def focus6(event):
    addr_field.focus_set()
# endregion

# region Clear all the entry boxes


def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_field.delete(0, END)
    addr_field.delete(0, END)
# endregion

# region Insert data from Window to Excel file


def insert():
    if(name_field.get() == "" and
       course_field.get() == "" and
       sem_field.get() == "" and
       form_no_field.get() == "" and
       contact_no_field.get() == "" and
       email_field.get() == "" and
       addr_field.get() == ""):

        print("empty input")

    else:

        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row+1, column=1).value = name_field.get()
        sheet.cell(row=current_row+1, column=2).value = course_field.get()
        sheet.cell(row=current_row+1, column=3).value = sem_field.get()
        sheet.cell(row=current_row+1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row+1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row+1, column=6).value = email_field.get()
        sheet.cell(row=current_row+1, column=7).value = addr_field.get()

        wb.save('data.xlsx')

        name_field.focus_set()

        clear()
# endregion

# region main
if __name__ == '__main__':
    label_pos = 0
    control_pos = 1
    box_pad = 100

    root = Tk()
    root.configure(background='light green')
    root.geometry("500x300")
    root.title('Application form')

    # region Labels
    heading = Label(root, text="Form", bg="light green")
    name = Label(root, text="Name", bg="light green")
    course = Label(root, text="Course", bg='light green')
    sem = Label(root, text='Semester', bg='light green')
    form_no = Label(root, text='Form No.', bg='light green')
    contact_no = Label(root, text='Contact No.', bg='light green')
    email = Label(root, text='Email ID', bg='light green')
    addr = Label(root, text='Address', bg='light green')
    # endregion

    # region Entries
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_field = Entry(root)
    addr_field = Entry(root)
    # endregion

    # region Binding
    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_field.bind("<Return>", focus6)

    # endregion

    # region Button
    excel()
    submit = Button(root, text="Submit", fg="black", bg="red", command=insert)
    # endregion

    # region Packing
    heading.grid(row=0, column=control_pos)
    name.grid(row=1, column=label_pos)
    course.grid(row=2, column=label_pos)
    sem.grid(row=3, column=label_pos)
    form_no.grid(row=4, column=label_pos)
    contact_no.grid(row=5, column=label_pos)
    email.grid(row=6, column=label_pos)
    addr.grid(row=7, column=label_pos)

    name_field.grid(row=1, column=control_pos, ipadx=box_pad)
    course_field.grid(row=2, column=control_pos, ipadx=box_pad)
    sem_field.grid(row=3, column=control_pos, ipadx=box_pad)
    form_no_field.grid(row=4, column=control_pos, ipadx=box_pad)
    contact_no_field.grid(row=5, column=control_pos, ipadx=box_pad)
    email_field.grid(row=6, column=control_pos, ipadx=box_pad)
    addr_field.grid(row=7, column=control_pos, ipadx=box_pad)

    submit.grid(row=8, column=control_pos)
    # endregion

    root.mainloop()
# endregion
